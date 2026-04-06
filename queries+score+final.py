import streamlit as st
import pandas as pd
import re
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# Regex pattern to capture numbers inside square brackets
pattern = r"\[(\d+(?:\.\d+)?)\]"

def extract_scores(cell):
    if pd.isna(cell):
        return []
    if isinstance(cell, list):
        scores = []
        for item in cell:
            matches = re.findall(pattern, str(item))
            scores.extend(matches)
        return scores
    else:
        return re.findall(pattern, str(cell))

def process_excel(uploaded_file):
    df = pd.read_excel(uploaded_file, sheet_name="Sheet1")

    # Identify question columns (skip "Name" and "Total Score")
    question_cols = [col for col in df.columns if col.startswith("Q")]

    # Build new dataframe
    new_df = pd.DataFrame()
    new_df["Name"] = df["Name"]

    for col in question_cols:
        scores_list = df[col].apply(extract_scores)
        max_scores = max(scores_list.apply(len)) if not scores_list.empty else 0
        max_scores = max(max_scores, 1)  # ensure at least one column
        for i in range(max_scores):
            new_df[f"{col}_Score{i+1}"] = scores_list.apply(lambda x: x[i] if i < len(x) else "")

    # Save to BytesIO
    output = BytesIO()
    new_df.to_excel(output, index=False)
    output.seek(0)

    # Post-process with openpyxl
    wb = load_workbook(output)
    ws = wb.active

    # First row: question titles merged across their score columns
    # Second row: all labeled "Score"
    for col in question_cols:
        score_cols = [c for c in new_df.columns if c.startswith(col + "_Score")]
        if score_cols:
            start_col = new_df.columns.get_loc(score_cols[0]) + 1
            end_col = new_df.columns.get_loc(score_cols[-1]) + 1
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            ws.cell(row=1, column=start_col).value = col
            ws.cell(row=1, column=start_col).alignment = Alignment(horizontal="center")
            for i, sc in enumerate(score_cols):
                col_idx = new_df.columns.get_loc(sc) + 1
                ws.cell(row=2, column=col_idx).value = "Score"
                ws.cell(row=2, column=col_idx).alignment = Alignment(horizontal="center")

    # Adjust "Name" column header
    ws.cell(row=1, column=1).value = "Name"
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    # Auto-adjust column widths safely
    for i in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(i)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=i, max_col=i):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # Save final workbook to BytesIO
    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    return final_output

# Streamlit UI
st.title("Excel Questionnaire Processor")

uploaded_file = st.file_uploader("Upload your responses.xlsx", type=["xlsx"])

if uploaded_file:
    st.success("File uploaded successfully!")
    processed_file = process_excel(uploaded_file)
    st.download_button(
        label="Download Processed Excel",
        data=processed_file,
        file_name="filtered_responses_scores_spanned.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
